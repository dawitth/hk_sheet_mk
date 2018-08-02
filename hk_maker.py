#!/usr/bin/env python3

import sys, re, os, os.path,subprocess
from tkinter import *
from tkinter.filedialog import askopenfilename
from datetime import datetime
from pandas import ExcelWriter
from openpyxl import load_workbook




def process():
    
    
    todays_date = datetime.strptime(todayss_date, '%d/%m/%y').date()

    # get the parent path of the exe
    exe_filepath= os.path.abspath(sys.argv[0])
    exe_filepath = os.path.dirname(exe_filepath)
    
    
   
    inhouse_filename_pdf = stayover_filename
    arrival_filename_pdf = arrival_filename
    
    pdfbox_app_path = exe_filepath + '/asset/pdfbox-app-2.0.11.jar'

    
 
    ### to avoid file access permission issue:
    ### copy pdfbox-app and the pdf files to the parent directory
    cmd = ['cp',pdfbox_app_path, exe_filepath]
    subprocess.call(cmd, shell=False)
    
    cmd = ['cp',inhouse_filename_pdf, exe_filepath]
    subprocess.call(cmd, shell=False)
    
    cmd = ['cp',arrival_filename_pdf, exe_filepath]
    subprocess.call(cmd, shell=False)
    
    ### get the copied file directory
    curr_inhouse_pdf = exe_filepath + "/" + os.path.basename(inhouse_filename_pdf)
    curr_arrival_pdf = exe_filepath + "/" + os.path.basename(arrival_filename_pdf)
    curr_pdfbox_path = exe_filepath + "/" + os.path.basename(pdfbox_app_path)
  
    
    ### extract text from the PDF
    
    cmd = ['java',"-jar", curr_pdfbox_path,"ExtractText",curr_inhouse_pdf]
    subprocess.call(cmd, shell=False)
    
    cmd = ['java',"-jar",curr_pdfbox_path,"ExtractText",curr_arrival_pdf]
    subprocess.call(cmd, shell=False)
    
    
    

    # generate path for txt file based on the pdf file name and exe path
    in_txt = os.path.basename(curr_inhouse_pdf)
    in_txt = in_txt[0:-3]
    in_txt = exe_filepath + "/" + in_txt + "txt"

    arrival_txt = os.path.basename(curr_arrival_pdf)
    arrival_txt = arrival_txt[0:-3]
    arrival_txt = exe_filepath + "/" + arrival_txt + "txt"
    
    #get data from the extracted file
    
    
    f_inhouse = open(in_txt, "r")
    f_arrival = open(arrival_txt, "r")
    

### remove copied files

    if os.path.exists(curr_inhouse_pdf):
        os.remove(curr_inhouse_pdf)
    if os.path.exists(curr_arrival_pdf):
        os.remove(curr_arrival_pdf)
    if os.path.exists(curr_pdfbox_path):
        os.remove(curr_pdfbox_path)
    if os.path.exists(in_txt):
        os.remove(in_txt)
    if os.path.exists(arrival_txt):
        os.remove(arrival_txt)
    


    ### populate room list with room numbers of the hotel
    room_list = []
    room_list += [[105],[106],[107],[108]]
    for x in range(211,250):
        tmp = []
        tmp.append(x)
        room_list.append(tmp)
    
    for x in range(351,391):
        tmp = []
        tmp.append(x)
        room_list.append(tmp)
    
    
   


    arrival_lines = list(f_arrival)
    lines = list(f_inhouse)

    f_inhouse.close()


    
    arrival_date = []
    departure_date = []
    room_numbers = []
    count  = 0;

    arrivals_arrival_date = []
    arrivals_room_numbers = []
    arrivals_count  = 0;
    

    arrival_lines#organizing the data by room number, arrival date and departure date
    #add a list with all the room numbres and test against it
    for i in range(len(lines)):
        match_room = re.search(r'^[123]\d{2}$',lines[i])
        match_date = re.search(r'^(^[0-3]\d\/\d{2}\/\d{2})$', lines[i])
        if match_room:
            #print(lines[i])
            room_numbers.append(lines[i])
            count +=1
        
        if match_date:
           
            date_obj = datetime.strptime(lines[i].rstrip(), '%d/%m/%y').date()
            
            if count == 0:
                departure_date.append(date_obj)
            else:
                arrival_date.append(date_obj)
                count -=1





    # processing the data to give S/O, C/O, A, B, C
    # processing stayovers

    #organizing the data by room number, arrival date and departure date
    #add a list with all the room numbres and test against it
    for i in range(len(arrival_lines)):
        match_room = re.search(r'^[123]\d{2}$',arrival_lines[i])
        match_date = re.search(r'^(^[0-3]\d\/\d{2}\/\d{2})$', arrival_lines[i])
        if match_room:
            #print(lines[i])
            arrivals_room_numbers.append(arrival_lines[i])
            arrivals_count +=1
        

        if match_date:
            
            date_obj = datetime.strptime(arrival_lines[i].rstrip(), '%d/%m/%y').date()
            
            if arrivals_count is not 0:
                arrivals_arrival_date.append(date_obj)
                arrivals_count -=1


    final_arrival_rooms = []
    for i in range(len(arrivals_arrival_date)):
        diffrence  = arrivals_arrival_date[i] - todays_date
      
        if diffrence.days is 0:
            final_arrival_rooms.append(int(arrivals_room_numbers[i]))

    final_data = room_list




    # assign 
    diffrence = 0
    for i in range(len(room_numbers)):
        room_data = []
        stayed_nights = todays_date - arrival_date[i]
        stayed_nights = stayed_nights.days
        departing_today = departure_date[i] - todays_date
        
        if departing_today.days is 0:
            room_data.append("C/O")
        elif departing_today.days < 0:
            room_data.append("x0")
        else:
            a_or_c = stayed_nights%3;
            
            if a_or_c is 0:
                room_data.append("S/O")
                room_data.append("A")
            else:
                room_data.append("S/O")
                room_data.append("C")
                    
        
        for x in final_data:
            
            if x[0] == int(room_numbers[i]):
                x +=(room_data)




    for i in range(len(final_data)):
       
        if final_data[i][0] in final_arrival_rooms and len(final_data[i]) < 3:
            final_data[i].append("R")


    for i in final_data:
        if len(i) == 1:
            i.append(" ")
            i.append(" ")
        if i[1] == "R":
            i.pop()
            i.append(" ")
            i.append("R")
        print(i)
    # save the final data to the spread sheet



    hk_filename = exe_filepath + '/asset/hksheet.xlsx'


    wb2 = load_workbook(hk_filename)

    ws = wb2['3rd_floor_1st_part']

    for i in range(len(final_data)):
        if len(final_data[i]) < 3:
            final_data[i].append("")


    for i in range(1, 151):
        cell = 'D' + str(i)
        rm_num = ws[cell].value
        for x in range(len(final_data)):
            
            if rm_num == final_data[x][0]:
                cell_e = 'E' + str(i)
                cell_f = 'F' + str(i)
                ws[cell_e].value = final_data[x][1]
                ws[cell_f].value = final_data[x][2]



    save_folder_path = exe_filepath + '/alances.xlsx'

    wb2.save(save_folder_path)







def main():
    
    global arrival_filename
    global stayover_filename
    global todays_date
    global save_folder_path
    
    root = Tk()
    
    ### gets arrival pdf directory
    def get_arr_dir(event):
        global arrival_filename
        
        ftypes = [('PDF',"*.pdf"),('All files',"*.*"),('TEXT',"*.txt")]
        root.fileName = askopenfilename(filetypes = ftypes)
        arrival_filename = root.fileName
        return
    
    ### gets stay over pdf directory
    def get_stay_dir(event):
        global stayover_filename
        
        ftypes = [('PDF',"*.pdf"),('All files',"*.*"),('TEXT',"*.txt")]
        root.fileName = askopenfilename(filetypes = ftypes)
        stayover_filename = root.fileName
        return
    
    ### gets date entered by the user
    def get_date(event):
        global todayss_date
        global save_folder_path

        todayss_date = dateEntry.get()
        
        process()
    
    
    
    Label(root, text="").grid(row=0, column=6)
    Label(root, text="Housekeeping task sheet", width=60).grid(row=1, column = 6)

    ### arrival button
    arrivalBtn = Button(root,   text="+  Arrival list")
    arrivalBtn.grid(row=3, column=6)
    arrivalBtn.bind("<Button-1>", get_arr_dir)

    ### stayover button
    stayoverBtn = Button(root,   text="+  Stayover list")
    stayoverBtn.grid(row=4, column=6)
    stayoverBtn.bind("<Button-1>", get_stay_dir)
    
    ### date entry field
    Label(root, text="Enter arrival date below as dd/mm/yy").grid(row=5, column=6)
    dateEntry = Entry(root, width=20)
    dateEntry.grid(row=6, column=6)
    
    ### process and save button
    processBtn = Button(root, text="Process and save")
    processBtn.grid(row=7, column=6)
    processBtn.bind("<Button-1>", get_date)
    
    
    
    Label(root, text="").grid(row=8, column=6)
    
    
    
    root.mainloop()







if __name__ == '__main__':
    main();
